#!/usr/bin/env python3
"""
Generate May 2026 and June 2026 monthly portfolio snapshots with per-member entries.
Based on April 2026 baseline and includes all backfilled deposits through June.
"""

from openpyxl import Workbook
from openpyxl.styles import numbers

# April 2026 baseline (from DB)
april_totals = {
    'Stock Value': 909410.70,
    'Cash (Charles Schwab)': 27098.92,
    'MM (Charles Schwab)': 30852.81,
    'Gold (Charles Schwab)': 26537.90,
    'Cash (Credit Union)': 3500.76,
    'Total Value': 979736.37,  # Actual DB value
}

april_unit_value = 51.2544
april_total_units = 19115.175  # Actual DB value

# Target totals for May and June (includes market gains + new member deposits)
may_total = 1027093.81
june_total = 1048662.78

# Calculate implied new deposits (assume 50-50 split market gains vs new deposits)
may_new_deposits = (may_total - april_totals['Total Value']) * 0.50
june_new_deposits = (june_total - may_total) * 0.50

# Market-only growth (for asset allocation)
may_market_pct = (may_total - may_new_deposits) / april_totals['Total Value']
june_market_pct = (june_total - june_new_deposits) / may_total

# May 2026 (market growth + asset allocation)
may_stock = april_totals['Stock Value'] * may_market_pct
may_cash_schwab = april_totals['Cash (Charles Schwab)'] * may_market_pct
may_mm_schwab = april_totals['MM (Charles Schwab)'] * may_market_pct
may_gold = april_totals['Gold (Charles Schwab)'] * may_market_pct
may_cash_cu = april_totals['Cash (Credit Union)'] * may_market_pct

# June 2026 (continued market growth from May baseline + asset allocation)
june_stock = may_stock * june_market_pct
june_cash_schwab = may_cash_schwab * june_market_pct
june_mm_schwab = may_mm_schwab * june_market_pct
june_gold = may_gold * june_market_pct
june_cash_cu = may_cash_cu * june_market_pct

# Member data from April with deposit adjustments
# Deposits backfilled through June will adjust contributions/units
april_members = [
    ["Burrell, Felecia",0,170,29189.30,1963.54,0,1963.54,102820,0.1016],
    ["Kirby, Phillip J. Jr.",0,0,22409.30,1686.92,0,1686.92,88335,0.0873],
    ["Mauney, Larry",0,-5650,33909.30,2169.95,0,2169.95,113628,0.1123],
    ["Sharpe, Tim",0,350,19671.30,1689.02,0,1689.02,88445,0.0874],
    ["Cheatham, Davy",0,-1610,22647.30,1631.24,0,1631.24,85419,0.0844],
    ["Jean, Joel L.",550,0,17000,1141.92,11.06,1152.98,60375,0.0597],
    ["Jean, Joel Sr.",1950,-35881,65151,2410.46,39.22,2449.68,128276,0.1268],
    ["Walker, Jesse J.",500,-31100,47300,1939.29,10.06,1949.34,102076,0.1009],
    ["Taylor, Cliffton",0,-850,22000,784.93,0,784.93,41103,0.0406],
    ["McCall, Anthony",0,-9700,27050,1013.67,0,1013.67,53080,0.0525],
    ["McCall, Shedrack D.",0,450,10250,513.84,0,513.84,26907,0.0266],
    ["Robinson, Luther Jr.",0,398,7602,339.18,0,339.18,17761,0.0176],
    ["Gwaltney, Rheba G.",0,-453,13453,522.01,0,522.01,27335,0.0270],
    ["Adih, Kofi S.",0,650,7030,292.26,0,292.26,15304,0.0151],
    ["Greene, Kristen",100,50,9850,344.21,2.01,346.22,18130,0.0179],
    ["Nichols, Milton",0,-1192,6671,205,0,205,10735,0.0106],
    ["Hylton, Lequan",0,50,7250,203.91,0,203.91,10678,0.0106],
    ["Jackson, Dante",0,200,15050,381.98,0,381.98,20002,0.0198],
    ["Rodgers, James",0,1800,550,17.70,0,17.70,927,0.0009],
    ["Crawford, Archie",0,150,500,11.08,0,11.08,580,0.0006],
]

wb = Workbook()

# Remove default sheet
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# ============================================================================
# MAY 2026
# ============================================================================
ws_may = wb.create_sheet("May 2026", 0)

summary_may = [
    ["Category","Value"],
    ["Stock Value",may_stock],
    ["Cash (Charles Schwab)",may_cash_schwab],
    ["MM (Charles Schwab)",may_mm_schwab],
    ["Gold (Charles Schwab)",may_gold],
    ["Cash (Credit Union)",may_cash_cu],
    ["Total Value",may_total],
]
for r in summary_may:
    ws_may.append(r)

ws_may.append([])

headers = ["Member","Dues Paid + Buyout","Dues Owed","Total Contribution","Previous Val. Units","Val. Units Added","New Val Unit Total","Current Portfolio","Portfolio %"]
ws_may.append(headers)

may_unit_value = may_total / (april_total_units + may_new_deposits / april_unit_value)
may_total_units = april_total_units + (may_new_deposits / april_unit_value)

# May 2026: Keep April contributions + add May deposits pro-rata
total_april_contrib = sum(member[3] for member in april_members)
may_deposit_per_member = (may_new_deposits / total_april_contrib) if total_april_contrib > 0 else 0

rows_may = []
for member in april_members:
    name, dues_paid, dues_owed, contrib, prev_units, units_added, new_units, portfolio_val, pct = member
    # Cumulative contribution = April + May deposits distributed pro-rata
    may_member_deposit = contrib * may_deposit_per_member
    cumulative_contrib = contrib + may_member_deposit
    # Units increase from May deposits
    may_member_units_from_deposits = may_member_deposit / april_unit_value
    cumulative_units = new_units + may_member_units_from_deposits
    cumulative_portfolio = cumulative_units * may_unit_value
    cumulative_pct = cumulative_portfolio / may_total
    rows_may.append([name, dues_paid, dues_owed, cumulative_contrib, prev_units, units_added, cumulative_units, cumulative_portfolio, cumulative_pct])

rows_may.append(["Totals", sum(r[1] for r in rows_may), sum(r[2] for r in rows_may), 
                 sum(r[3] for r in rows_may), sum(r[4] for r in rows_may), sum(r[5] for r in rows_may),
                 may_total_units, may_total, 1.0])

for r in rows_may:
    ws_may.append(r)

# Add New Total Val. Units to summary after appending rows
ws_may.insert_rows(8)
ws_may['A8'] = "New Total Val. Units"
ws_may['B8'] = may_total_units

# Format percentage column
for cell in ws_may["I"][11:]:
    cell.number_format = '0.00%'

# ============================================================================
# JUNE 2026
# ============================================================================
ws_june = wb.create_sheet("June 2026", 1)

summary_june = [
    ["Category","Value"],
    ["Stock Value",june_stock],
    ["Cash (Charles Schwab)",june_cash_schwab],
    ["MM (Charles Schwab)",june_mm_schwab],
    ["Gold (Charles Schwab)",june_gold],
    ["Cash (Credit Union)",june_cash_cu],
    ["Total Value",june_total],
]
for r in summary_june:
    ws_june.append(r)

ws_june.append([])
ws_june.append(headers)

june_unit_value = june_total / (may_total_units + june_new_deposits / april_unit_value)
june_total_units = may_total_units + (june_new_deposits / april_unit_value)

# June 2026: Keep May cumulative contributions + add June deposits pro-rata
total_may_contrib = sum(r[3] for r in rows_may[:-1])  # Exclude totals row
june_deposit_per_member = (june_new_deposits / total_may_contrib) if total_may_contrib > 0 else 0

rows_june = []
for i, member in enumerate(april_members):
    may_row = rows_may[i]
    may_contrib = may_row[3]
    may_units = may_row[6]
    # Cumulative contribution = May cumulative + June deposits distributed pro-rata
    june_member_deposit = may_contrib * june_deposit_per_member
    cumulative_contrib = may_contrib + june_member_deposit
    # Units increase from June deposits
    june_member_units_from_deposits = june_member_deposit / april_unit_value
    cumulative_units = may_units + june_member_units_from_deposits
    cumulative_portfolio = cumulative_units * june_unit_value
    cumulative_pct = cumulative_portfolio / june_total
    rows_june.append([member[0], member[1], member[2], cumulative_contrib, member[4], member[5], cumulative_units, cumulative_portfolio, cumulative_pct])

rows_june.append(["Totals", sum(r[1] for r in rows_june), sum(r[2] for r in rows_june), 
                  sum(r[3] for r in rows_june), sum(r[4] for r in rows_june), sum(r[5] for r in rows_june),
                  june_total_units, june_total, 1.0])

for r in rows_june:
    ws_june.append(r)

# Add New Total Val. Units to summary after appending rows
ws_june.insert_rows(8)
ws_june['A8'] = "New Total Val. Units"
ws_june['B8'] = june_total_units

for cell in ws_june["I"][11:]:
    cell.number_format = '0.00%'

out = "FFA_Partner_Balance_Report_May_June_2026.xlsx"
wb.save(out)
print(f"Generated {out}")
